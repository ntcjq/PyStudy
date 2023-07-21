import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


def generateExcel():
    wb = Workbook()
    # 创建了一个用户信息表并放在第一位
    ws = wb.create_sheet('用户信息表', 0)
    #设置A1的列宽
    ws.column_dimensions["A"].width = 30
    #设置A1的行高
    ws.row_dimensions[1].height = 20

    #指定位置赋值
    # ws['A1'] = 42
    # ws['A2'] = datetime.datetime.now()
    # ws.cell(row=4, column=2, value=10)

    #顺着一行一行的赋值
    ws.append(['name', 'age', 'gender'])
    ws.append(['cjq', '30', '1'])
    ws.append(['ssn', '29', '2'])

    wb.save("file/sample.xlsx")
    print("Excel生成结束")


def readExcel():
    try:
        wb = load_workbook('file/sample.xlsx', read_only=True, data_only=True)
        ws = wb['用户信息表']
        for row in ws.rows:  # 获取每一行的数据
            for data in row:  # 获取每一行中单元格的数据
                print(data.value, end="\t")  # 打印单元格的值
            print()

    except FileNotFoundError as e:
        print('指定的文件无法打开.')
    except IOError as e:
        print('读写文件时出现错误.')
    print('程序执行结束.')


if __name__ == '__main__':
    generateExcel()
    readExcel()